import json
import os
import re
import subprocess
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import ProtectedError
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

from .models import MenuItem, Order, OrderItem

EXPECTED_XLSX_HEADERS = ["菜品名稱", "菜品價格", "過敏原", "菜品介紹"]
MAX_CHAT_MESSAGES = 12
MAX_CART_ACTION_QUANTITY = 20


def serialize_item(item):
    return {
        "id": item.id,
        "name": item.name,
        "description": item.description,
        "price": str(item.price),
        "allergens": item.allergens,
    }


def serialize_order(order):
    return {
        "id": order.id,
        "created_at": order.created_at.isoformat(),
        "total_price": str(order.total_price),
        "items": [
            {
                "id": order_item.id,
                "menu_item_id": order_item.menu_item_id,
                "menu_item_name": order_item.menu_item.name,
                "quantity": order_item.quantity,
                "unit_price": str(order_item.unit_price),
                "line_total": str(order_item.line_total),
            }
            for order_item in order.items.select_related("menu_item").all()
        ],
    }


def parse_payload(request):
    if not request.body:
        return {}

    try:
        return json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError("JSON 格式不正確。") from exc


def validate_menu_item_payload(payload):
    name = str(payload.get("name", "")).strip()
    description = str(payload.get("description", "")).strip()
    allergens = str(payload.get("allergens", "")).strip()
    raw_price = payload.get("price", "")

    if not name:
        raise ValueError("菜品名稱為必填。")

    try:
        price = Decimal(str(raw_price))
    except (InvalidOperation, TypeError) as exc:
        raise ValueError("菜品價格必須是有效數字。") from exc

    if price < 0:
        raise ValueError("菜品價格不能小於 0。")

    return {
        "name": name,
        "description": description,
        "price": price,
        "allergens": allergens,
    }


def validate_order_payload(payload):
    raw_items = payload.get("items")

    if not isinstance(raw_items, list) or not raw_items:
        raise ValueError("訂單至少要有一個品項。")

    validated_items = []

    for index, raw_item in enumerate(raw_items, start=1):
        if not isinstance(raw_item, dict):
            raise ValueError(f"第 {index} 筆訂單項目格式不正確。")

        menu_item_id = raw_item.get("menu_item_id")
        quantity = raw_item.get("quantity")

        try:
            menu_item_id = int(menu_item_id)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"第 {index} 筆菜品 ID 必須是整數。") from exc

        try:
            quantity = int(quantity)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"第 {index} 筆數量必須是整數。") from exc

        if quantity <= 0:
            raise ValueError(f"第 {index} 筆數量必須大於 0。")

        validated_items.append(
            {
                "menu_item_id": menu_item_id,
                "quantity": quantity,
            }
        )

    return validated_items


def validate_chat_messages(payload):
    raw_messages = payload.get("messages")
    if not isinstance(raw_messages, list) or not raw_messages:
        raise ValueError("請至少提供一則聊天訊息。")

    validated = []

    for index, message in enumerate(raw_messages, start=1):
        if not isinstance(message, dict):
            raise ValueError(f"第 {index} 則訊息格式不正確。")

        role = str(message.get("role", "")).strip()
        content = str(message.get("content", "")).strip()

        if role not in {"user", "assistant"}:
            raise ValueError(f"第 {index} 則訊息角色只接受 user 或 assistant。")

        if not content:
            raise ValueError(f"第 {index} 則訊息內容不可為空白。")

        validated.append({"role": role, "content": content})

    return validated[-MAX_CHAT_MESSAGES:]


def validate_cart_snapshot(payload):
    raw_cart = payload.get("cart", [])

    if raw_cart in (None, ""):
        return []

    if not isinstance(raw_cart, list):
        raise ValueError("購物車資料格式不正確。")

    validated = []
    for index, item in enumerate(raw_cart, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"第 {index} 筆購物車資料格式不正確。")

        try:
            menu_item_id = int(item.get("menu_item_id"))
            quantity = int(item.get("quantity"))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"第 {index} 筆購物車資料格式不正確。") from exc

        if quantity <= 0:
            continue

        validated.append({"menu_item_id": menu_item_id, "quantity": quantity})

    return validated


def build_menu_context():
    items = MenuItem.objects.all()
    if not items:
        return "目前沒有菜單資料可供推薦。"

    lines = []
    for item in items:
        allergens = item.allergens or "無"
        description = item.description or "尚無介紹"
        lines.append(
            f"- ID {item.id}｜{item.name}｜價格 NT$ {item.price}｜過敏原：{allergens}｜介紹：{description}"
        )
    return "\n".join(lines)


def build_cart_context(cart_snapshot):
    if not cart_snapshot:
        return "目前購物車是空的。"

    items_by_id = {item.id: item for item in MenuItem.objects.filter(id__in=[x["menu_item_id"] for x in cart_snapshot])}
    lines = []
    for entry in cart_snapshot:
        menu_item = items_by_id.get(entry["menu_item_id"])
        if menu_item is None:
            lines.append(f"- ID {entry['menu_item_id']}｜數量 {entry['quantity']}（此品項目前已不在菜單中）")
            continue
        lines.append(f"- ID {menu_item.id}｜{menu_item.name}｜數量 {entry['quantity']}")
    return "\n".join(lines)


def extract_text_response(response_data):
    candidates = response_data.get("candidates", [])
    for candidate in candidates:
        content = candidate.get("content", {})
        for part in content.get("parts", []):
            text = str(part.get("text", "")).strip()
            if text:
                return text

    prompt_feedback = response_data.get("promptFeedback")
    if prompt_feedback:
        raise ValueError(f"模型沒有回傳內容：{prompt_feedback}")

    raise ValueError("模型沒有回傳可用的文字內容。")


def build_llm_contents(messages):
    contents = []
    for message in messages:
        role = "model" if message["role"] == "assistant" else "user"
        contents.append(
            {
                "role": role,
                "parts": [{"text": message["content"]}],
            }
        )
    return contents


def build_llm_payload(model, messages, instruction_text):
    contents = build_llm_contents(messages)

    if model.startswith("gemma-"):
        return {
            "contents": [
                {
                    "role": "user",
                    "parts": [{"text": instruction_text}],
                },
                *contents,
            ],
            "generationConfig": {
                "temperature": 0.2,
                "maxOutputTokens": 1024,
            },
        }

    return {
        "system_instruction": {
            "parts": [{"text": instruction_text}],
        },
        "contents": contents,
        "generationConfig": {
            "temperature": 0.2,
            "maxOutputTokens": 1024,
        },
    }


def build_cli_prompt(messages, instruction_text):
    lines = [instruction_text, "", "對話紀錄："]

    for message in messages:
        speaker = "助理" if message["role"] == "assistant" else "使用者"
        lines.append(f"{speaker}: {message['content']}")

    lines.append("")
    lines.append("請只輸出符合格式的 JSON。")
    return "\n".join(lines)


def extract_text_from_json_node(node):
    if isinstance(node, str):
        value = node.strip()
        return value or None

    if isinstance(node, dict):
        for key in ("response", "text", "content", "message", "result"):
            value = node.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()

        for value in node.values():
            extracted = extract_text_from_json_node(value)
            if extracted:
                return extracted

    if isinstance(node, list):
        for item in node:
            extracted = extract_text_from_json_node(item)
            if extracted:
                return extracted

    return None


def extract_cli_text_response(raw_output):
    output = raw_output.strip()
    if not output:
        raise ValueError("Gemini CLI 沒有回傳任何內容。")

    try:
        payload = json.loads(output)
    except json.JSONDecodeError:
        return output

    if isinstance(payload, dict) and payload.get("error"):
        raise ValueError(f"Gemini CLI 回傳錯誤：{payload['error']}")

    extracted = extract_text_from_json_node(payload)
    if extracted:
        return extracted

    raise ValueError("Gemini CLI 沒有回傳可解析的文字內容。")


def call_gemini_cli(model, messages, cart_snapshot):
    command = os.environ.get("GEMINI_CLI_COMMAND", "gemini").strip() or "gemini"
    prompt = build_cli_prompt(messages, build_chat_instruction(cart_snapshot))
    cli_env = os.environ.copy()
    cli_env.pop("GEMINI_API_KEY", None)

    completed = subprocess.run(
        [command, "-m", model, "-p", prompt, "--output-format", "json"],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="ignore",
        timeout=90,
        cwd="/app",
        env=cli_env,
        check=False,
    )

    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout).strip() or f"exit code {completed.returncode}"
        raise RuntimeError(f"Gemini CLI 執行失敗：{detail}")

    return extract_cli_text_response(completed.stdout)


def call_gemini_api(model, api_key, messages, cart_snapshot):
    llm_payload = build_llm_payload(model, messages, build_chat_instruction(cart_snapshot))
    endpoint = (
        f"{GEMINI_API_BASE}/{urlparse.quote(model, safe='')}:generateContent"
        f"?key={urlparse.quote(api_key, safe='')}"
    )

    req = urlrequest.Request(
        endpoint,
        data=json.dumps(llm_payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urlrequest.urlopen(req, timeout=45) as response:
            response_data = json.loads(response.read().decode("utf-8"))
    except urlerror.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"Gemini API 回傳錯誤：{detail or exc.reason}") from exc
    except urlerror.URLError as exc:
        raise RuntimeError(f"無法連線到 Gemini API：{exc.reason}") from exc

    return extract_text_response(response_data)


def extract_json_object(text):
    text = text.strip()
    if not text:
        raise ValueError("模型沒有回傳內容。")

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or start >= end:
        raise ValueError("模型沒有回傳 JSON。")

    return json.loads(text[start : end + 1])


def normalize_chat_actions(raw_actions):
    if not isinstance(raw_actions, list):
        return []

    valid_menu_ids = set(MenuItem.objects.values_list("id", flat=True))
    normalized = []

    for action in raw_actions:
        if not isinstance(action, dict):
            continue

        action_type = str(action.get("type", "")).strip().lower()

        if action_type in {"clear_cart", "clear"}:
            normalized.append({"type": "clear_cart"})
            continue

        if action_type in {"remove_item", "remove"}:
            try:
                menu_item_id = int(action.get("menu_item_id"))
            except (TypeError, ValueError):
                continue

            if menu_item_id in valid_menu_ids:
                normalized.append(
                    {
                        "type": "remove_item",
                        "menu_item_id": menu_item_id,
                    }
                )
            continue

        if action_type in {"set_quantity", "add_item", "update_item"}:
            try:
                menu_item_id = int(action.get("menu_item_id"))
                quantity = int(action.get("quantity"))
            except (TypeError, ValueError):
                continue

            if menu_item_id not in valid_menu_ids:
                continue

            quantity = max(0, min(quantity, MAX_CART_ACTION_QUANTITY))
            normalized.append(
                {
                    "type": "set_quantity",
                    "menu_item_id": menu_item_id,
                    "quantity": quantity,
                }
            )

    return normalized[:10]


def parse_chat_response(response_text):
    try:
        response_json = extract_json_object(response_text)
    except (ValueError, json.JSONDecodeError):
        return {
            "reply": response_text.strip(),
            "actions": [],
        }

    reply = str(response_json.get("reply", "")).strip()
    if not reply:
        reply = "已依照您的需求更新。"

    return {
        "reply": reply,
        "actions": normalize_chat_actions(response_json.get("actions", [])),
    }


def actions_mutate_cart(actions):
    return any(
        action.get("type") in {"set_quantity", "remove_item", "clear_cart"}
        for action in actions
        if isinstance(action, dict)
    )


def reply_claims_cart_mutation(reply):
    normalized = str(reply or "")
    mutation_markers = (
        "已加入",
        "已加到",
        "加入購物車",
        "已更新",
        "更新為",
        "已為您更新",
        "已移除",
        "已刪除",
        "已取消",
        "已清空",
        "清空購物車",
        "已將",
    )
    return any(marker in normalized for marker in mutation_markers)


def bind_reply_to_actions(parsed_response, messages, cart_snapshot):
    actions = parsed_response.get("actions", [])
    if actions_mutate_cart(actions):
        parsed_response["reply"] = build_action_reply(actions)
        return parsed_response

    fallback_actions = infer_actions_from_latest_message(messages, cart_snapshot)
    if fallback_actions:
        return {
            "reply": build_action_reply(fallback_actions),
            "actions": fallback_actions,
        }

    if reply_claims_cart_mutation(parsed_response.get("reply", "")):
        return {
            "reply": "我還沒有更新購物車。請說明品項名稱與數量，或回覆「我要」承接上一個明確品項。",
            "actions": [],
        }

    return parsed_response


def find_latest_user_message(messages):
    for message in reversed(messages):
        if message["role"] == "user":
            return message["content"].strip()
    return ""


def find_previous_user_message(messages):
    found_latest = False
    for message in reversed(messages):
        if message["role"] != "user":
            continue
        if not found_latest:
            found_latest = True
            continue
        return message["content"].strip()
    return ""


def build_latest_user_turn(messages):
    latest_message = find_latest_user_message(messages)
    if not latest_message:
        return messages[-1:]
    return [{"role": "user", "content": latest_message}]


def is_unhelpful_model_reply(reply):
    normalized = str(reply or "").strip()
    if not normalized:
        return False

    refusal_markers = (
        "抱歉",
        "我無法",
        "無法隨機",
        "不能隨機",
        "我不能",
        "我只能",
        "沒有辦法",
    )
    return any(marker in normalized for marker in refusal_markers)


def is_confirmation_message(text):
    normalized = normalize_lookup_text(text)
    return normalized in {
        "對",
        "對的",
        "是",
        "是的",
        "好",
        "好啊",
        "可以",
        "沒錯",
        "正確",
        "就這個",
        "就這樣",
        "這個",
        "那個",
        "ok",
        "okay",
    }


def is_contextual_order_message(text):
    normalized = normalize_lookup_text(text)
    if is_confirmation_message(text):
        return True

    order_phrases = (
        "我要",
        "要這個",
        "點這個",
        "選這個",
        "來一份",
        "來一杯",
        "來一個",
        "好來一份",
        "好來一杯",
        "好我要",
        "就這個",
        "就來這個",
        "那就這個",
    )
    if any(phrase in normalized for phrase in order_phrases):
        return True

    return (
        any(verb in normalized for verb in ("來", "要", "點", "加"))
        and any(unit in normalized for unit in ("份", "杯", "個", "碗", "盤", "瓶", "罐"))
    )


def is_quantity_only_message(text):
    normalized = normalize_lookup_text(text)
    quantity = extract_quantity_from_text(text, default=None)
    if quantity is None:
        return False

    quantity_words = ("零", "〇", "一", "二", "兩", "三", "四", "五", "六", "七", "八", "九", "十")
    unit_words = ("份", "杯", "個", "碗", "盤", "瓶", "罐", "條", "顆", "籠", "客")
    command_words = ("改", "改成", "改為", "換", "要", "來", "加", "點", "變成")

    has_quantity = any(char.isdigit() for char in normalized) or any(word in normalized for word in quantity_words)
    has_unit = any(word in normalized for word in unit_words)
    has_contextual_command = any(word in normalized for word in command_words)
    return has_quantity and (has_unit or has_contextual_command) and len(normalized) <= 8


def build_action_reply(actions):
    if not actions:
        return ""

    item_names = {
        item.id: item.name
        for item in MenuItem.objects.filter(
            id__in=[action.get("menu_item_id") for action in actions if action.get("menu_item_id")]
        )
    }

    if len(actions) == 1:
        action = actions[0]
        if action["type"] == "clear_cart":
            return "已為您清空購物車。"
        if action["type"] == "remove_item":
            name = item_names.get(action["menu_item_id"], f"ID {action['menu_item_id']}")
            return f"已從購物車移除 {name}。"
        if action["type"] == "set_quantity":
            name = item_names.get(action["menu_item_id"], f"ID {action['menu_item_id']}")
            return f"已將 {name} 更新為 {action['quantity']} 份。"

    return "已依照您的需求更新購物車。"


def normalize_lookup_text(text):
    normalized = []
    for char in str(text).lower():
        if char.isascii() and char.isalnum():
            normalized.append(char)
            continue
        if "\u4e00" <= char <= "\u9fff":
            normalized.append(char)
    return "".join(normalized)


def parse_chinese_number(token):
    digits = {
        "零": 0,
        "〇": 0,
        "一": 1,
        "二": 2,
        "兩": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
    }

    token = token.strip()
    if not token:
        return None

    if token.isdigit():
        return int(token)

    if token == "十":
        return 10

    if "十" in token:
        parts = token.split("十", 1)
        tens = digits.get(parts[0], 1 if parts[0] == "" else None)
        ones = digits.get(parts[1], 0 if parts[1] == "" else None)
        if tens is None or ones is None:
            return None
        return tens * 10 + ones

    total = 0
    for char in token:
        value = digits.get(char)
        if value is None:
            return None
        total = total * 10 + value
    return total


def extract_quantity_from_text(text, default=1):
    patterns = [
        r"(?:加|來|點|要|給我|幫我加|幫我點|再來|再加|改|改成|改為|換|變成)\s*([0-9零〇一二兩三四五六七八九十]+)",
        r"([0-9零〇一二兩三四五六七八九十]+)\s*(?:份|個|杯|碗|盤|瓶|罐|條|顆|籠|客|份量)",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        quantity = parse_chinese_number(match.group(1))
        if quantity is not None:
            return max(0, min(quantity, MAX_CART_ACTION_QUANTITY))

    return default


def build_menu_alias_map(item_ids=None):
    queryset = MenuItem.objects.all()
    if item_ids is not None:
        queryset = queryset.filter(id__in=item_ids)

    alias_candidates = {}
    items_by_id = {}

    for item in queryset:
        normalized_name = normalize_lookup_text(item.name)
        items_by_id[item.id] = item
        aliases = {normalized_name}

        max_suffix_length = min(len(normalized_name), 8)
        for length in range(3, max_suffix_length + 1):
            aliases.add(normalized_name[-length:])

        for start in range(1, min(len(normalized_name), 4)):
            alias = normalized_name[start:]
            if len(alias) >= 4:
                aliases.add(alias)

        for alias in aliases:
            alias_candidates.setdefault(alias, set()).add(item.id)

    unique_aliases = {
        alias: next(iter(item_ids))
        for alias, item_ids in alias_candidates.items()
        if len(item_ids) == 1 and len(alias) >= 3
    }

    return items_by_id, unique_aliases


def find_item_ids_in_text(text, item_ids=None):
    normalized_text = normalize_lookup_text(text)
    if not normalized_text:
        return []

    _, alias_map = build_menu_alias_map(item_ids)
    matches = []
    consumed = normalized_text

    for alias, item_id in sorted(alias_map.items(), key=lambda item: len(item[0]), reverse=True):
        if alias in consumed and item_id not in matches:
            matches.append(item_id)
            consumed = consumed.replace(alias, " ", 1)

    return matches


def find_referenced_menu_item_ids(text, item_ids=None):
    valid_ids = set(MenuItem.objects.values_list("id", flat=True))
    if item_ids is not None:
        valid_ids &= {int(item_id) for item_id in item_ids}

    matches = []
    for match in re.finditer(r"\bID\s*(\d+)\b", str(text), re.IGNORECASE):
        item_id = int(match.group(1))
        if item_id in valid_ids and item_id not in matches:
            matches.append(item_id)

    for item_id in find_item_ids_in_text(text, item_ids):
        if item_id not in matches:
            matches.append(item_id)

    return matches


def find_contextual_fragment_item_ids(text, item_ids):
    normalized_text = normalize_lookup_text(text)
    if len(normalized_text) < 2:
        return []

    matches = []
    for item in MenuItem.objects.filter(id__in=item_ids):
        haystack = normalize_lookup_text(f"{item.name} {item.description or ''}")
        if normalized_text in haystack:
            matches.append(item.id)

    return matches


def find_recent_referenced_menu_item_ids(messages):
    latest_user_skipped = False
    for message in reversed(messages):
        if message["role"] == "user" and not latest_user_skipped:
            latest_user_skipped = True
            continue

        referenced_ids = find_referenced_menu_item_ids(message["content"])
        if referenced_ids:
            return referenced_ids

    return []


def infer_actions_from_latest_message(messages, cart_snapshot):
    latest_message = find_latest_user_message(messages)
    if not latest_message:
        return []

    add_keywords = ("加", "加入", "來", "點", "我要", "給我", "幫我加", "幫我點", "再來", "再加")
    update_keywords = ("改成", "改為", "改做", "改", "換成", "換")
    delete_keywords = ("刪", "移除", "取消", "不要", "拿掉")
    clear_keywords = ("清空", "全部刪掉", "全部移除", "清掉")
    reference_keywords = ("剛剛", "剛才", "上一個", "上個", "那個", "這個")
    info_keywords = ("多少錢", "價格", "價錢", "過敏原", "介紹", "推薦", "有什麼", "是什麼", "好吃嗎")
    cart_item_ids = [item["menu_item_id"] for item in cart_snapshot]
    mentioned_cart_items = find_item_ids_in_text(latest_message, cart_item_ids)
    mentioned_menu_items = find_referenced_menu_item_ids(latest_message)
    previous_user_message = find_previous_user_message(messages)

    if is_quantity_only_message(latest_message):
        quantity = extract_quantity_from_text(latest_message, default=1)
        recent_item_ids = find_recent_referenced_menu_item_ids(messages)
        target_item_ids = recent_item_ids or (cart_item_ids if len(cart_item_ids) == 1 else [])
        if target_item_ids:
            return normalize_chat_actions(
                [
                    {
                        "type": "set_quantity",
                        "menu_item_id": target_item_ids[0],
                        "quantity": quantity,
                    }
                ]
            )

    if is_contextual_order_message(latest_message):
        recent_item_ids = find_recent_referenced_menu_item_ids(messages)
        if recent_item_ids:
            return normalize_chat_actions(
                [
                    {
                        "type": "set_quantity",
                        "menu_item_id": recent_item_ids[0],
                        "quantity": extract_quantity_from_text(latest_message, default=1),
                    }
                ]
            )

    if any(keyword in latest_message for keyword in info_keywords):
        return []

    if any(keyword in latest_message for keyword in clear_keywords) and not any(
        blocked_keyword in latest_message for blocked_keyword in ("菜單", "菜品", "全部資料")
    ):
        return [{"type": "clear_cart"}]

    id_match = re.search(r"ID\s*(\d+)", latest_message, re.IGNORECASE)
    if id_match and (
        is_contextual_order_message(latest_message)
        or any(keyword in latest_message for keyword in update_keywords + add_keywords)
    ):
        return normalize_chat_actions(
            [
                {
                    "type": "set_quantity",
                    "menu_item_id": int(id_match.group(1)),
                    "quantity": extract_quantity_from_text(latest_message, default=1),
                }
            ]
        )

    if id_match and any(keyword in latest_message for keyword in delete_keywords):
        return normalize_chat_actions(
            [
                {
                    "type": "remove_item",
                    "menu_item_id": int(id_match.group(1)),
                }
            ]
        )

    if mentioned_cart_items and any(keyword in latest_message for keyword in delete_keywords):
        return normalize_chat_actions(
            [{"type": "remove_item", "menu_item_id": menu_item_id} for menu_item_id in mentioned_cart_items]
        )

    if (
        len(cart_snapshot) == 1
        and any(keyword in latest_message for keyword in delete_keywords)
        and any(keyword in latest_message for keyword in reference_keywords)
    ):
        return normalize_chat_actions(
            [
                {
                    "type": "remove_item",
                    "menu_item_id": cart_snapshot[0]["menu_item_id"],
                }
            ]
        )

    order_context_keywords = add_keywords + ("喝", "一杯", "杯", "冰", "冷", "飲料", "想喝", "幫我找")
    if (
        mentioned_menu_items
        and previous_user_message
        and any(keyword in previous_user_message for keyword in order_context_keywords)
    ):
        return normalize_chat_actions(
            [
                {
                    "type": "set_quantity",
                    "menu_item_id": menu_item_id,
                    "quantity": extract_quantity_from_text(previous_user_message, default=1),
                }
                for menu_item_id in mentioned_menu_items
            ]
        )

    if mentioned_menu_items and any(keyword in latest_message for keyword in update_keywords + add_keywords):
        if len(mentioned_menu_items) == 1:
            quantity = extract_quantity_from_text(latest_message, default=1)
            return normalize_chat_actions(
                [
                    {
                        "type": "set_quantity",
                        "menu_item_id": mentioned_menu_items[0],
                        "quantity": quantity,
                    }
                ]
            )

        return normalize_chat_actions(
            [
                {
                    "type": "set_quantity",
                    "menu_item_id": menu_item_id,
                    "quantity": 1,
                }
                for menu_item_id in mentioned_menu_items
            ]
        )

    return []


def build_chat_instruction(cart_snapshot):
    return (
        "你是智慧點餐系統中的點餐助理。請使用繁體中文回答，回覆要直接、清楚、簡短。"
        "你只負責理解使用者意圖與產生候選 actions，後端會驗證並執行 actions。"
        "你只能根據提供的菜單資料推薦、說明價格、提醒過敏原，或幫客人操作購物車。"
        "你不能直接送出訂單，只能透過 actions 控制購物車。"
        "你必須只回傳 JSON，不要加 markdown、不要加說明文字。"
        'JSON 格式固定為 {"reply":"文字回覆","actions":[...]}。'
        "actions 可用的類型只有："
        'set_quantity（需要 menu_item_id 與 quantity，可新增或改數量，quantity=0 代表移除）、'
        'remove_item（需要 menu_item_id）、clear_cart（不需要其他欄位）。'
        "如果 reply 說已加入、已更新、已移除或已清空，actions 必須包含對應動作。"
        "如果 actions 是空陣列，reply 絕對不能說已經更新購物車。"
        "如果使用者只是詢問菜單資訊，不要亂動購物車，actions 請回傳空陣列。"
        "如果使用者要求加入、刪除、修改購物車，你必須優先使用菜單中的正確 ID。"
        "如果使用者只說好、對、我要、來一份、兩份、改兩杯、就這個，請根據最近一次明確提到或推薦的菜品判斷。"
        "如果最近上下文只有一個明確菜品，省略句要套用到該菜品。"
        "如果最近上下文有多個候選品項，模糊詞只能在最近候選品項中比對，不可以擴大到全菜單。"
        "若菜單沒有對應品項，reply 要明確說明找不到，actions 保持空陣列。"
        f"\n\n目前菜單：\n{build_menu_context()}"
        f"\n\n目前購物車：\n{build_cart_context(cart_snapshot)}"
    )


def parse_menu_items_xlsx(uploaded_file):
    if not uploaded_file.name.lower().endswith(".xlsx"):
        raise ValueError("只接受 .xlsx 檔案。")

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row = next(
        worksheet.iter_rows(min_row=1, max_row=1, max_col=4, values_only=True),
        None,
    )
    normalized_headers = [str(value or "").strip() for value in (header_row or ())]

    if normalized_headers != EXPECTED_XLSX_HEADERS:
        raise ValueError("Excel 欄位格式不符合要求，請使用 test.xlsx 的欄位順序。")

    parsed_items = []

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=2, max_col=4, values_only=True),
        start=2,
    ):
        name, price, allergens, description = row

        if all(value in (None, "") for value in (name, price, allergens, description)):
            continue

        try:
            parsed_items.append(
                validate_menu_item_payload(
                    {
                        "name": name,
                        "price": price,
                        "allergens": allergens or "",
                        "description": description or "",
                    }
                )
            )
        except ValueError as exc:
            raise ValueError(f"第 {row_index} 列資料有誤：{exc}") from exc

    if not parsed_items:
        raise ValueError("Excel 內沒有可匯入的菜品資料。")

    return parsed_items


@csrf_exempt
def menu_items(request):
    if request.method == "GET":
        items = [serialize_item(item) for item in MenuItem.objects.all()]
        return JsonResponse(items, safe=False)

    if request.method == "POST":
        try:
            payload = validate_menu_item_payload(parse_payload(request))
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)

        item = MenuItem.objects.create(**payload)
        return JsonResponse(serialize_item(item), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
def chat_with_gpt(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    model = os.environ.get("GEMINI_MODEL", "gemma-3-27b-it").strip() or "gemma-3-27b-it"

    if not api_key:
        return JsonResponse(
            {"error": "後端尚未設定 GEMINI_API_KEY，請先在 Docker 環境變數中設定。"},
            status=503,
        )

    try:
        payload = parse_payload(request)
        messages = validate_chat_messages(payload)
        cart_snapshot = validate_cart_snapshot(payload)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    shortcut_actions = infer_actions_from_latest_message(messages, cart_snapshot)
    if shortcut_actions:
        return JsonResponse(
            {
                "reply": build_action_reply(shortcut_actions),
                "actions": shortcut_actions,
                "model": model,
            }
        )

    llm_payload = build_llm_payload(model, messages, build_chat_instruction(cart_snapshot))

    endpoint = (
        f"{GEMINI_API_BASE}/{urlparse.quote(model, safe='')}:generateContent"
        f"?key={urlparse.quote(api_key, safe='')}"
    )

    req = urlrequest.Request(
        endpoint,
        data=json.dumps(llm_payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urlrequest.urlopen(req, timeout=45) as response:
            response_data = json.loads(response.read().decode("utf-8"))
    except urlerror.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        return JsonResponse(
            {"error": f"Gemini API 回傳錯誤：{detail or exc.reason}"},
            status=502,
        )
    except urlerror.URLError as exc:
        return JsonResponse({"error": f"無法連線到 Gemini API：{exc.reason}"}, status=502)
    except Exception as exc:
        return JsonResponse({"error": f"Gemini 呼叫失敗：{exc}"}, status=500)

    try:
        response_text = extract_text_response(response_data)
        parsed_response = parse_chat_response(response_text)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=502)

    if not parsed_response["actions"]:
        fallback_actions = infer_actions_from_latest_message(messages, cart_snapshot)
        if fallback_actions:
            parsed_response["actions"] = fallback_actions
            if not parsed_response["reply"] or parsed_response["reply"].replace("?", "").strip() == "":
                parsed_response["reply"] = build_action_reply(fallback_actions)

    if not parsed_response["reply"]:
        parsed_response["reply"] = build_action_reply(parsed_response["actions"]) or "請問還需要其他調整嗎？"

    return JsonResponse(
        {
            "reply": parsed_response["reply"],
            "actions": parsed_response["actions"],
            "model": model,
        }
    )


@csrf_exempt
def import_menu_items_xlsx(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    uploaded_file = request.FILES.get("file")
    if uploaded_file is None:
        return JsonResponse({"error": "請上傳 xlsx 檔案。"}, status=400)

    try:
        parsed_items = parse_menu_items_xlsx(uploaded_file)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    with transaction.atomic():
        OrderItem.objects.all().delete()
        Order.objects.all().delete()
        MenuItem.objects.all()._raw_delete(MenuItem.objects.db)
        MenuItem.objects.bulk_create([MenuItem(**item) for item in parsed_items])

    return JsonResponse(
        {
            "message": f"Excel 匯入完成，舊菜單與舊訂單已清空，已覆蓋 {len(parsed_items)} 筆菜品資料。",
            "count": len(parsed_items),
        }
    )


@csrf_exempt
def menu_item_detail(request, item_id):
    try:
        item = MenuItem.objects.get(id=item_id)
    except MenuItem.DoesNotExist:
        return JsonResponse({"error": "找不到指定菜品。"}, status=404)

    if request.method == "GET":
        return JsonResponse(serialize_item(item))

    if request.method == "PUT":
        try:
            payload = validate_menu_item_payload(parse_payload(request))
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)

        for field, value in payload.items():
            setattr(item, field, value)
        item.save()
        return JsonResponse(serialize_item(item))

    if request.method == "DELETE":
        try:
            item.delete()
        except ProtectedError:
            return JsonResponse({"error": "此菜品已有訂單紀錄，暫時無法刪除。"}, status=400)
        return JsonResponse({"message": "菜品已刪除。"})

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
def orders(request):
    if request.method == "GET":
        order_list = [
            serialize_order(order)
            for order in Order.objects.prefetch_related("items__menu_item").all()[:20]
        ]
        return JsonResponse(order_list, safe=False)

    if request.method == "POST":
        try:
            validated_items = validate_order_payload(parse_payload(request))
        except ValueError as exc:
            return JsonResponse({"error": str(exc)}, status=400)

        menu_item_ids = [item["menu_item_id"] for item in validated_items]
        menu_items_by_id = {
            item.id: item for item in MenuItem.objects.filter(id__in=menu_item_ids)
        }

        missing_ids = sorted(set(menu_item_ids) - set(menu_items_by_id))
        if missing_ids:
            return JsonResponse(
                {"error": f"找不到菜品 ID：{', '.join(str(item_id) for item_id in missing_ids)}"},
                status=404,
            )

        with transaction.atomic():
            total_price = Decimal("0.00")
            normalized_items = []

            for item in validated_items:
                menu_item = menu_items_by_id[item["menu_item_id"]]
                quantity = item["quantity"]
                unit_price = menu_item.price
                line_total = unit_price * quantity
                total_price += line_total
                normalized_items.append(
                    {
                        "menu_item": menu_item,
                        "quantity": quantity,
                        "unit_price": unit_price,
                        "line_total": line_total,
                    }
                )

            order = Order.objects.create(total_price=total_price)

            OrderItem.objects.bulk_create(
                [
                    OrderItem(
                        order=order,
                        menu_item=item["menu_item"],
                        quantity=item["quantity"],
                        unit_price=item["unit_price"],
                        line_total=item["line_total"],
                    )
                    for item in normalized_items
                ]
            )

        order = Order.objects.prefetch_related("items__menu_item").get(id=order.id)
        return JsonResponse(serialize_order(order), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
def order_detail(request, order_id):
    try:
        order = Order.objects.prefetch_related("items__menu_item").get(id=order_id)
    except Order.DoesNotExist:
        return JsonResponse({"error": "找不到指定訂單。"}, status=404)

    if request.method == "GET":
        return JsonResponse(serialize_order(order))

    if request.method == "DELETE":
        order.delete()
        return JsonResponse({"message": "訂單已刪除。"})

    return JsonResponse({"error": "Method not allowed"}, status=405)


@csrf_exempt
def chat_with_llm(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        payload = parse_payload(request)
        messages = validate_chat_messages(payload)
        cart_snapshot = validate_cart_snapshot(payload)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    model = os.environ.get("GEMINI_CLI_MODEL", "gemini-2.5-flash").strip() or "gemini-2.5-flash"

    shortcut_actions = infer_actions_from_latest_message(messages, cart_snapshot)
    if shortcut_actions:
        return JsonResponse(
            {
                "reply": build_action_reply(shortcut_actions),
                "actions": shortcut_actions,
                "model": model,
            }
        )

    try:
        if provider == "gemini-cli":
            response_text = call_gemini_cli(model, messages, cart_snapshot)
        else:
            api_key = os.environ.get("GEMINI_API_KEY", "").strip()
            if not api_key:
                return JsonResponse(
                    {"error": "後端尚未設定 GEMINI_API_KEY，請先在 Docker 環境變數中設定。"},
                    status=503,
                )
            response_text = call_gemini_api(model, api_key, messages, cart_snapshot)
    except RuntimeError as exc:
        return JsonResponse({"error": str(exc)}, status=502)
    except subprocess.TimeoutExpired:
        return JsonResponse({"error": "Gemini CLI 執行逾時，請確認 CLI 已完成登入並可正常回應。"}, status=504)
    except Exception as exc:
        return JsonResponse({"error": f"Gemini 發生例外：{exc}"}, status=500)

    try:
        parsed_response = parse_chat_response(response_text)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=502)

    if not parsed_response["actions"]:
        fallback_actions = infer_actions_from_latest_message(messages, cart_snapshot)
        if fallback_actions:
            parsed_response["actions"] = fallback_actions
            if not parsed_response["reply"] or parsed_response["reply"].replace("?", "").strip() == "":
                parsed_response["reply"] = build_action_reply(fallback_actions)

    if not parsed_response["reply"]:
        parsed_response["reply"] = build_action_reply(parsed_response["actions"]) or "已完成回覆。"

    return JsonResponse(
        {
            "reply": parsed_response["reply"],
            "actions": parsed_response["actions"],
            "model": model,
        }
    )


@csrf_exempt
def chat_with_cli(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        payload = parse_payload(request)
        messages = validate_chat_messages(payload)
        cart_snapshot = validate_cart_snapshot(payload)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    model = os.environ.get("GEMINI_CLI_MODEL", "gemini-2.5-flash").strip() or "gemini-2.5-flash"

    shortcut_actions = infer_actions_from_latest_message(messages, cart_snapshot)
    if shortcut_actions:
        return JsonResponse(
            {
                "reply": build_action_reply(shortcut_actions),
                "actions": shortcut_actions,
                "model": model,
            }
        )

    try:
        response_text = call_gemini_cli(model, messages, cart_snapshot)
    except RuntimeError as exc:
        return JsonResponse({"error": str(exc)}, status=502)
    except subprocess.TimeoutExpired:
        return JsonResponse({"error": "Gemini CLI 執行逾時，請確認 CLI 已完成登入並可正常回應。"}, status=504)
    except Exception as exc:
        return JsonResponse({"error": f"Gemini 發生例外：{exc}"}, status=500)

    try:
        parsed_response = parse_chat_response(response_text)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=502)

    if not parsed_response["actions"]:
        fallback_actions = infer_actions_from_latest_message(messages, cart_snapshot)
        if fallback_actions:
            parsed_response["actions"] = fallback_actions
            if not parsed_response["reply"] or parsed_response["reply"].replace("?", "").strip() == "":
                parsed_response["reply"] = build_action_reply(fallback_actions)

    if not parsed_response["reply"]:
        parsed_response["reply"] = build_action_reply(parsed_response["actions"]) or "已完成回覆。"

    return JsonResponse(
        {
            "reply": parsed_response["reply"],
            "actions": parsed_response["actions"],
            "model": model,
        }
    )


chat_with_llm = chat_with_cli
chat_with_gpt = chat_with_cli


def get_menu_items_in_order(item_ids):
    items_by_id = {
        item.id: item
        for item in MenuItem.objects.filter(id__in=item_ids)
    }
    return [items_by_id[item_id] for item_id in item_ids if item_id in items_by_id]


def format_menu_item_brief(item):
    return f"{item.name}（NT$ {item.price}）"


def format_menu_item_detail(item):
    allergens = item.allergens or "未標示過敏原"
    description = item.description or "目前沒有更多介紹"
    return f"ID {item.id}｜{item.name}｜NT$ {item.price}｜過敏原：{allergens}｜{description}"


def build_contextual_menu_reply(messages):
    latest_message = find_latest_user_message(messages)
    if not latest_message:
        return None

    recent_item_ids = find_recent_referenced_menu_item_ids(messages)
    if not recent_item_ids:
        return None

    matched_item_ids = find_referenced_menu_item_ids(latest_message, recent_item_ids)
    if not matched_item_ids:
        matched_item_ids = find_contextual_fragment_item_ids(latest_message, recent_item_ids)
    if not matched_item_ids:
        return None

    matched_items = get_menu_items_in_order(matched_item_ids[:5])
    if not matched_items:
        return None

    if len(matched_items) == 1:
        item = matched_items[0]
        return {
            "reply": f"您剛剛提到的選項中，符合的是 {format_menu_item_detail(item)}。如果要點，請回覆「我要」或「來一份」。",
            "actions": [],
        }

    return {
        "reply": "您剛剛提到的選項中，符合的是：\n"
        + "\n".join(f"- {format_menu_item_detail(item)}" for item in matched_items)
        + "\n請回覆完整品名或說「我要 ID 編號」。",
        "actions": [],
    }


def extract_budget_amount(text):
    amounts = []
    for match in re.finditer(r"(?:NT\$?\s*)?(\d{2,5})(?:\s*(?:元|塊|台幣|ntd|NTD))?", text):
        try:
            amounts.append(Decimal(match.group(1)))
        except InvalidOperation:
            continue
    return max(amounts) if amounts else None


def decimal_to_cents(value):
    return int(Decimal(value) * 100)


def choose_budget_combo(menu_items, target_amount, max_items=10):
    target_cents = decimal_to_cents(target_amount)
    states = {0: []}

    for item in menu_items:
        item_cents = decimal_to_cents(item.price)
        if item_cents <= 0:
            continue

        next_states = dict(states)
        for current_cents, combo in states.items():
            if len(combo) >= max_items:
                continue

            next_cents = current_cents + item_cents
            if next_cents > target_cents:
                continue

            next_combo = combo + [item]
            if next_cents not in next_states or len(next_combo) < len(next_states[next_cents]):
                next_states[next_cents] = next_combo

        states = next_states

    best_cents = max(states)
    return states[best_cents], Decimal(best_cents) / Decimal("100")


def build_budget_combo_reply(latest_message):
    combo_keywords = ("湊", "預算", "套餐", "組合", "搭配", "配一套", "配餐")
    if not any(keyword in latest_message for keyword in combo_keywords):
        return None

    target_amount = extract_budget_amount(latest_message)
    if target_amount is None:
        return None

    menu_items = list(MenuItem.objects.all().order_by("-price", "id"))
    if not menu_items:
        return {
            "reply": "目前沒有可搭配的菜品。",
            "actions": [],
        }

    chosen_items, total_price = choose_budget_combo(menu_items, target_amount)

    if not chosen_items:
        cheapest_item = min(menu_items, key=lambda item: item.price)
        return {
            "reply": (
                f"目前最便宜的品項是 {cheapest_item.name}（NT$ {cheapest_item.price}），"
                f"已超過您指定的 NT$ {target_amount:.2f}，所以先不加入購物車。"
            ),
            "actions": [],
        }

    remaining = target_amount - total_price
    budget_note = ""
    if remaining > 0:
        budget_note = f"，距離 NT$ {target_amount:.2f} 還差 NT$ {remaining:.2f}"

    return {
        "reply": (
            "我幫您搭配："
            + "、".join(format_menu_item_brief(item) for item in chosen_items)
            + f"，合計 NT$ {total_price:.2f}{budget_note}，已加入購物車。"
        ),
        "actions": [
            {
                "type": "set_quantity",
                "menu_item_id": item.id,
                "quantity": 1,
            }
            for item in chosen_items
        ],
    }


def build_local_menu_reply(latest_message, cart_snapshot):
    price_keywords = ("多少錢", "價格", "價錢", "多少")
    allergen_keywords = ("過敏原", "會過敏", "有什麼不能吃")
    description_keywords = ("介紹", "是什麼", "內容", "說明")
    recommendation_keywords = ("推薦", "有什麼", "想吃", "吃什麼", "來點", "套餐")

    budget_combo_reply = build_budget_combo_reply(latest_message)
    if budget_combo_reply:
        return budget_combo_reply

    mentioned_item_ids = find_item_ids_in_text(latest_message)
    mentioned_items = get_menu_items_in_order(mentioned_item_ids[:5])

    if cart_snapshot and any(keyword in latest_message for keyword in ("總金額", "合計", "一共", "多少錢")):
        items_by_id = {
            item.id: item for item in MenuItem.objects.filter(id__in=[entry["menu_item_id"] for entry in cart_snapshot])
        }
        total_price = Decimal("0.00")
        for entry in cart_snapshot:
            menu_item = items_by_id.get(entry["menu_item_id"])
            if menu_item is not None:
                total_price += menu_item.price * entry["quantity"]
        if total_price > 0:
            return {
                "reply": f"目前購物車總金額為 NT$ {total_price:.2f}。",
                "actions": [],
            }

    if mentioned_items and any(keyword in latest_message for keyword in price_keywords):
        if len(mentioned_items) == 1:
            item = mentioned_items[0]
            return {
                "reply": f"{item.name} 的價格是 NT$ {item.price}。",
                "actions": [],
            }
        return {
            "reply": "；".join(f"{item.name} NT$ {item.price}" for item in mentioned_items),
            "actions": [],
        }

    if mentioned_items and any(keyword in latest_message for keyword in allergen_keywords):
        if len(mentioned_items) == 1:
            item = mentioned_items[0]
            allergens = item.allergens or "未標示過敏原"
            return {
                "reply": f"{item.name} 的過敏原資訊：{allergens}。",
                "actions": [],
            }
        return {
            "reply": "；".join(
                f"{item.name}：{item.allergens or '未標示過敏原'}"
                for item in mentioned_items
            ),
            "actions": [],
        }

    if mentioned_items and any(keyword in latest_message for keyword in description_keywords):
        item = mentioned_items[0]
        description = item.description or "目前沒有更多介紹。"
        return {
            "reply": f"{item.name}：{description}",
            "actions": [],
        }

    if any(keyword in latest_message for keyword in recommendation_keywords):
        all_items = list(MenuItem.objects.all())
        if not all_items:
            return {
                "reply": "目前沒有可推薦的菜品。",
                "actions": [],
            }

        keyword_groups = [
            ("飲料", ("飲料", "喝", "茶", "奶茶", "咖啡", "冷泡", "果汁", "汽水")),
            ("酒", ("酒", "啤酒", "紅酒", "白酒", "highball", "沙瓦")),
            ("飯類", ("飯", "炒飯", "燴飯", "丼", "粥")),
            ("麵類", ("麵", "拉麵", "拌麵", "湯麵", "義大利麵")),
            ("餃類", ("餃", "水餃", "煎餃", "鍋貼")),
            ("餅類", ("餅", "捲餅", "蔥油餅", "蛋餅")),
        ]

        normalized_message = normalize_lookup_text(latest_message)
        candidates = all_items

        for _, keywords in keyword_groups:
            if not any(normalize_lookup_text(keyword) in normalized_message for keyword in keywords):
                continue

            filtered = []
            for item in all_items:
                haystack = normalize_lookup_text(f"{item.name} {item.description or ''}")
                if any(normalize_lookup_text(keyword) in haystack for keyword in keywords):
                    filtered.append(item)

            if filtered:
                candidates = filtered
                break

        suggestions = candidates[:3]
        return {
            "reply": "推薦您：" + "、".join(format_menu_item_brief(item) for item in suggestions) + "。",
            "actions": [],
        }

    if any(keyword in latest_message for keyword in ("加", "加入", "刪", "移除", "不要", "清空")) and not mentioned_items:
        return {
            "reply": "我目前找不到對應的菜品，請直接輸入菜名或使用菜單上的名稱。",
            "actions": [],
        }

    return {
        "reply": "您可以直接說菜名與數量，例如「幫我加兩份蔥香牛肉捲餅」或「推薦一個飯類」。",
        "actions": [],
    }


def build_zero_config_chat_response(messages, cart_snapshot, model):
    shortcut_actions = infer_actions_from_latest_message(messages, cart_snapshot)
    if shortcut_actions:
        return {
            "reply": build_action_reply(shortcut_actions),
            "actions": shortcut_actions,
            "model": model,
        }

    local_response = build_local_menu_reply(find_latest_user_message(messages), cart_snapshot)
    if local_response["actions"]:
        return {
            "reply": local_response["reply"],
            "actions": local_response["actions"],
            "model": "local-rules",
        }

    contextual_response = build_contextual_menu_reply(messages)
    if contextual_response:
        return {
            "reply": contextual_response["reply"],
            "actions": contextual_response["actions"],
            "model": "local-rules",
        }

    try:
        response_text = call_gemini_cli(model, build_latest_user_turn(messages), cart_snapshot)
        parsed_response = parse_chat_response(response_text)
    except (RuntimeError, subprocess.TimeoutExpired, ValueError, json.JSONDecodeError, OSError, Exception):
        parsed_response = build_local_menu_reply(find_latest_user_message(messages), cart_snapshot)
        parsed_response = bind_reply_to_actions(parsed_response, messages, cart_snapshot)
        return {
            "reply": parsed_response["reply"],
            "actions": parsed_response["actions"],
            "model": "local-rules",
        }

    if not parsed_response["actions"]:
        fallback_actions = infer_actions_from_latest_message(messages, cart_snapshot)
        if fallback_actions:
            parsed_response["actions"] = fallback_actions
            if not parsed_response["reply"]:
                parsed_response["reply"] = build_action_reply(fallback_actions)

    if not parsed_response["actions"]:
        local_response = build_local_menu_reply(find_latest_user_message(messages), cart_snapshot)
        if local_response["actions"]:
            parsed_response = local_response
        elif is_unhelpful_model_reply(parsed_response["reply"]):
            parsed_response = local_response

    if not parsed_response["reply"]:
        local_response = build_local_menu_reply(find_latest_user_message(messages), cart_snapshot)
        parsed_response["reply"] = local_response["reply"]
        if not parsed_response["actions"]:
            parsed_response["actions"] = local_response["actions"]

    parsed_response = bind_reply_to_actions(parsed_response, messages, cart_snapshot)

    return {
        "reply": parsed_response["reply"],
        "actions": parsed_response["actions"],
        "model": model,
    }


@csrf_exempt
def chat_with_zero_config(request):
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        payload = parse_payload(request)
        messages = validate_chat_messages(payload)
        cart_snapshot = validate_cart_snapshot(payload)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    model = os.environ.get("GEMINI_CLI_MODEL", "gemini-2.5-flash").strip() or "gemini-2.5-flash"
    response_payload = build_zero_config_chat_response(messages, cart_snapshot, model)
    return JsonResponse(response_payload)


chat_with_llm = chat_with_zero_config
chat_with_gpt = chat_with_zero_config
